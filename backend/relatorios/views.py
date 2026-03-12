import calendar
import re
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import requests
from django.conf import settings
from django.core.cache import cache
from django.shortcuts import render


SOLVED_OR_CLOSED_STATUSES = {5, 6}

OPEN_BACKLOG_STATUSES = {1, 2, 3, 4}
ASSIGNED_STATUSES = {2, 3}
PENDING_STATUS = 4
RESOLVED_STATUS = 5
CLOSED_STATUS = 6


def normalize_value(value):
    if value in (None, "", "-"):
        return "-"

    if isinstance(value, list):
        return ", ".join(str(item) for item in value if item) or "-"

    return str(value)




import io
from datetime import datetime

import requests
from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

try:
    from xhtml2pdf import pisa
    PDF_AVAILABLE = True
except Exception:
    PDF_AVAILABLE = False


GLPI_STATUS_MAP = {
    1: "Novo",
    2: "Processando (Atribuído)",
    3: "Processando (Planejado)",
    4: "Pendente",
    5: "Resolvido",
    6: "Fechado",
}


def glpi_headers(session_token=None):
    headers = {
        "Content-Type": "application/json",
    }

    if session_token:
        headers["Session-Token"] = session_token

    if getattr(settings, "GLPI_APP_TOKEN", ""):
        headers["App-Token"] = settings.GLPI_APP_TOKEN

    return headers


def build_base_context(request):
    return {
        "glpi_name": request.session.get("glpi_name", request.user.username),
        "glpi_profile": request.session.get("glpi_profile", "Sem perfil"),
    }


def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return None


def format_date(value):
    if not value:
        return "-"
    try:
        dt = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return value


def glpi_get_dropdown(endpoint, session_token, range_value="0-999"):
    url = f"{settings.GLPI_API_URL}/{endpoint}"
    try:
        response = requests.get(
            url,
            headers=glpi_headers(session_token),
            params={"range": range_value},
            verify=getattr(settings, "GLPI_VERIFY_SSL", True),
            timeout=30,
        )
        if response.status_code not in (200, 206):
            return []
        data = response.json()
        return data if isinstance(data, list) else []
    except requests.RequestException:
        return []


def glpi_search_users(session_token):
    """
    Retorna usuários para popular o filtro de técnico.
    """
    url = f"{settings.GLPI_API_URL}/search/User"
    users = []
    start = 0
    step = 100

    while True:
        params = {
            "reset": "reset",
            "forcedisplay[0]": 2,   # id
            "forcedisplay[1]": 1,   # login
            "forcedisplay[2]": 9,   # realname
            "forcedisplay[3]": 34,  # firstname
            "range": f"{start}-{start + step - 1}",
        }

        try:
            response = requests.get(
                url,
                headers=glpi_headers(session_token),
                params=params,
                verify=getattr(settings, "GLPI_VERIFY_SSL", True),
                timeout=30,
            )

            if response.status_code not in (200, 206):
                break

            payload = response.json()
            rows = payload.get("data", [])
            if not rows:
                break

            for row in rows:
                realname = (row.get("9") or "").strip()
                firstname = (row.get("34") or "").strip()
                display_name = f"{realname} {firstname}".strip() or row.get("1") or "Sem nome"
                users.append({
                    "id": row.get("2"),
                    "name": display_name,
                })

            total_count = payload.get("totalcount", 0)
            start += step
            if start >= total_count:
                break

        except requests.RequestException:
            break

    users.sort(key=lambda x: (x["name"] or "").lower())
    return users


def glpi_search_tickets(session_token, filters, start=0, limit=9999):
    """
    Busca chamados no GLPI via search/Ticket.
    Suporta paginação: start e limit definem o range da consulta ao GLPI.
    Retorna (chamados, totalcount).
    """
    url = f"{settings.GLPI_API_URL}/search/Ticket"
    end = start + limit - 1
    range_str = f"{start}-{end}"

    params = {
        "reset": "reset",
        "forcedisplay[0]": settings.GLPI_TICKET_SEARCH_ID_FIELD_ID,          # ID
        "forcedisplay[1]": settings.GLPI_TICKET_SEARCH_NAME_FIELD_ID,        # Título
        "forcedisplay[2]": settings.GLPI_TICKET_SEARCH_STATUS_FIELD_ID,      # Status
        "forcedisplay[3]": settings.GLPI_TICKET_SEARCH_DATE_FIELD_ID,        # Abertura
        "forcedisplay[4]": settings.GLPI_TICKET_SEARCH_CLOSEDATE_FIELD_ID,   # Fechamento
        "forcedisplay[5]": settings.GLPI_TICKET_SEARCH_ASSIGNEE_FIELD_ID,    # Técnico
        "forcedisplay[6]": settings.GLPI_TICKET_SEARCH_GROUP_FIELD_ID,       # Grupo
        "forcedisplay[7]": settings.GLPI_TICKET_SEARCH_LOCATION_FIELD_ID,    # Localização
        "forcedisplay[8]": settings.GLPI_TICKET_SEARCH_CATEGORY_FIELD_ID,    # Categoria
        "range": range_str,
    }

    criteria_index = 0

    if filters.get("data_inicio"):
        params[f"criteria[{criteria_index}][field]"] = settings.GLPI_TICKET_SEARCH_DATE_FIELD_ID
        params[f"criteria[{criteria_index}][searchtype]"] = "morethan"
        params[f"criteria[{criteria_index}][value]"] = filters["data_inicio"].strftime("%Y-%m-%d 00:00:00")
        criteria_index += 1

    if filters.get("data_fim"):
        params[f"criteria[{criteria_index}][field]"] = settings.GLPI_TICKET_SEARCH_DATE_FIELD_ID
        params[f"criteria[{criteria_index}][searchtype]"] = "lessthan"
        params[f"criteria[{criteria_index}][value]"] = filters["data_fim"].strftime("%Y-%m-%d 23:59:59")
        criteria_index += 1

    if filters.get("tecnico"):
        params[f"criteria[{criteria_index}][field]"] = settings.GLPI_TICKET_SEARCH_ASSIGNEE_FIELD_ID
        params[f"criteria[{criteria_index}][searchtype]"] = "equals"
        params[f"criteria[{criteria_index}][value]"] = filters["tecnico"]
        criteria_index += 1

    if filters.get("grupo"):
        params[f"criteria[{criteria_index}][field]"] = settings.GLPI_TICKET_SEARCH_GROUP_FIELD_ID
        params[f"criteria[{criteria_index}][searchtype]"] = "equals"
        params[f"criteria[{criteria_index}][value]"] = filters["grupo"]
        criteria_index += 1

    if filters.get("localizacao"):
        params[f"criteria[{criteria_index}][field]"] = settings.GLPI_TICKET_SEARCH_LOCATION_FIELD_ID
        params[f"criteria[{criteria_index}][searchtype]"] = "equals"
        params[f"criteria[{criteria_index}][value]"] = filters["localizacao"]
        criteria_index += 1

    if filters.get("categoria"):
        params[f"criteria[{criteria_index}][field]"] = settings.GLPI_TICKET_SEARCH_CATEGORY_FIELD_ID
        params[f"criteria[{criteria_index}][searchtype]"] = "equals"
        params[f"criteria[{criteria_index}][value]"] = filters["categoria"]
        criteria_index += 1

    try:
        response = requests.get(
            url,
            headers=glpi_headers(session_token),
            params=params,
            verify=getattr(settings, "GLPI_VERIFY_SSL", True),
            timeout=60,
        )

        if response.status_code not in (200, 206):
            return [], 0

        payload = response.json()
        rows = payload.get("data", [])
        totalcount = int(payload.get("totalcount", len(rows)))

        id_key = str(settings.GLPI_TICKET_SEARCH_ID_FIELD_ID)
        title_key = str(settings.GLPI_TICKET_SEARCH_NAME_FIELD_ID)
        status_key = str(settings.GLPI_TICKET_SEARCH_STATUS_FIELD_ID)
        date_key = str(settings.GLPI_TICKET_SEARCH_DATE_FIELD_ID)
        close_key = str(settings.GLPI_TICKET_SEARCH_CLOSEDATE_FIELD_ID)
        tech_key = str(settings.GLPI_TICKET_SEARCH_ASSIGNEE_FIELD_ID)
        group_key = str(settings.GLPI_TICKET_SEARCH_GROUP_FIELD_ID)
        location_key = str(settings.GLPI_TICKET_SEARCH_LOCATION_FIELD_ID)
        category_key = str(settings.GLPI_TICKET_SEARCH_CATEGORY_FIELD_ID)

        chamados = []
        for row in rows:
            raw_status = row.get(status_key)
            try:
                status_int = int(raw_status) if raw_status not in (None, "") else None
            except Exception:
                status_int = None

            chamados.append({
                "id": row.get(id_key),
                "titulo": normalize_value(row.get(title_key)),
                "tecnico": normalize_value(row.get(tech_key)),
                "grupo": normalize_value(row.get(group_key)),
                "localizacao": normalize_value(row.get(location_key)),
                "categoria": normalize_value(row.get(category_key)),
                "status": GLPI_STATUS_MAP.get(status_int, str(raw_status) if raw_status else "-"),
                "status_id": status_int,
                "data_abertura": format_date(row.get(date_key)),
                "data_fechamento": format_date(row.get(close_key)),
            })

        return chamados, totalcount

    except requests.RequestException:
        return [], 0


def build_summary(chamados, filters):
    total_chamados = len(chamados)
    chamados_fechados = sum(1 for c in chamados if c["status_id"] == 6)
    chamados_abertos = sum(1 for c in chamados if c["status_id"] in (1, 2, 3, 4, 5))

    tecnicos_set = {
        normalize_value(c["tecnico"])
        for c in chamados
        if normalize_value(c["tecnico"]) != "-"
    }

    grupos_set = {
        normalize_value(c["grupo"])
        for c in chamados
        if normalize_value(c["grupo"]) != "-"
    }

    return {
        "total_chamados": total_chamados,
        "chamados_fechados": chamados_fechados,
        "chamados_abertos": chamados_abertos,
        "total_tecnicos": len(tecnicos_set),
        "total_grupos": len(grupos_set),
    }

def export_excel(chamados):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatórios"

    headers = [
        "ID",
        "Título",
        "Técnico",
        "Grupo",
        "Localização",
        "Categoria",
        "Status",
        "Data abertura",
        "Data fechamento",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for chamado in chamados:
        ws.append([
            chamado["id"],
            chamado["titulo"],
            chamado["tecnico"],
            chamado["grupo"],
            chamado["localizacao"],
            chamado["categoria"],
            chamado["status"],
            chamado["data_abertura"],
            chamado["data_fechamento"],
        ])

    widths = {
        "A": 10,
        "B": 40,
        "C": 28,
        "D": 25,
        "E": 25,
        "F": 30,
        "G": 18,
        "H": 22,
        "I": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="relatorio_chamados.xlsx"'
    return response


def export_pdf(request, context):
    if not PDF_AVAILABLE:
        return HttpResponse(
            "Exportação PDF indisponível: instale xhtml2pdf (pip install xhtml2pdf)",
            status=501,
            content_type="text/plain",
        )

    html = render(request, "relatorios/relatorios_pdf.html", context).content.decode("utf-8")
    result = io.BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=result, encoding="utf-8")

    if pisa_status.err:
        return HttpResponse(
            "Erro ao gerar PDF.",
            status=500,
            content_type="text/plain",
        )

    response = HttpResponse(result.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="relatorio_chamados.pdf"'
    return response


